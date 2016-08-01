import warnings
import os
from flask import render_template, request, make_response, url_for, send_file, flash
from openpyxl import load_workbook
from openpyxl.writer.excel import save_workbook
from sqlalchemy import desc
from sqlalchemy.sql import func
from app import app
from app import db
from models import Video, Company, Company_List, Single_Report_View, Current_Month_Stats
from models import Video_Tag, TopCompany, CurrentMonth,VideoTopCompany
from models import Masterclass_Top_Companies, Month_Report, Channel_Reports, Tag,TagList
from models import Channel_Reports_Page1,Channel_Reports_Last_Month, month_table
from models import CHANNEL_YTD_Audience_Category,Last_Month,CHANNEL_YTD_Reports,Channel_last_month_Detail,channel_report_all_time_2
from models import channel_email_stats
from forms import  VideoEditForm , CompanyByNameForm, VideoByNameForm, CurrentMonthForm


@app.route('/')
@app.route('/index')
def index():
    return render_template('index.html')

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500

@app.route('/show_report_month', methods=['GET', 'POST'])
def report_month():

    monthData = CurrentMonth.query.one()
    print monthData

    form = CurrentMonthForm(obj=monthData)
    form.month_number.default = monthData.month_number
    form.month_year.default = monthData.month_year

    if form.validate_on_submit():

        form.populate_obj(obj=monthData)
        db.session.commit()
        flash('Change Completed!')


    return render_template('show_current_month.html', form = form)


@app.route('/search/video', methods=['GET', 'POST'])
def search_title():

    form = VideoByNameForm()
    results = None

    if form.validate_on_submit():
        title = '%' + form.title.data + '%'
        results = Video.query.filter(Video.V_Title.ilike(title.encode("utf-8"))).all()
        for r in results:
            print r
    return render_template('video_search.html',results=results,form=form)


@app.route('/edit/video/<key>', methods=['POST'])
def get_video(key):
    video = Video.query.get(key)
    form = VideoEditForm(request.form)
    form.populate_obj(video)

    return render_template('video_edit.html',form=form)


@app.route('/search/company', methods=['GET', 'POST'])
def search_company():

    form = CompanyByNameForm()
    results = None

    if form.validate_on_submit():
        company_name = '%' + form.CName.data + '%'
        results = Company_List.query.filter(Company_List.CName.ilike(company_name.encode("utf-8"))).all()
        for r in results:
            print r
    return render_template('company_search.html',results=results,form=form)

@app.route('/edit/company/<key>', methods=['POST'])
def get_company(key):
    #company = Company.query.get(key)
    form = VideoEditForm(request.form)
    #form.populate_obj(video)
    print 'ready to render template'

    return render_template('video_edit.html',form=form)

#
#  get video_id for single report
#
@app.route('/search/single_report', methods=['GET', 'POST'])
def get_single_report():

    #form = VideoByNameForm()
    results = Video.query.filter((Video.V_Type =='SINGLE') | (Video.V_Type =='MASTERCLASS'))\
            .order_by(Video.V_Title).all()

    return render_template('single_report_search.html',results=results)
#
#  produce single report
#
@app.route('/reports/single_report/<int:video_id>', methods=['POST'])
def produce_single_report(video_id):
    summary = {}

    results = Single_Report_View.query.filter_by(V_ID = video_id).order_by(Single_Report_View.month_id).all()

    # did we get reporting data? If not, complain
    if not results:
        return render_template('no_report_data.html')

    #get current month stats from database
    current_month_stats = Current_Month_Stats.query.all()

    # unable to retrieve the current reportng month/year, show error page
    if not current_month_stats:
        return render_template('error.html',error_message='Unable to retrieve the current reportng month/year, check database')

    current_month = current_month_stats[0].month_name # March, April .....
    current_year =  current_month_stats[0].month_year # 2015, 2016 .....
    current_month_number  = current_month_stats[0].month_number # 1 = Jan, 3 = Mar......
    current_month_id = current_month_stats[0].month_id # 14,15,16 primary key for Month_Reports table

    # is this a single or masterclass video?
    masterclass = True
    single = False
    if results[0].VType == 'SINGLE':
        masterclass = False
        single = True

    # pie chart needs a list of list
    point = []
    audience_profile = []

    #dictionary for report header
    header = {}

    #dictionary for audience profile in spreadsheet
    audience_profile_dict = {}
    if single:

        #----------------------------------
        # single report processing
        #----------------------------------

        # loop through results set to find current month

        for result_row in results:
           if result_row.SPeriod == current_month and result_row.SYear == current_year :

               summary['total_views'] = result_row.total_views
               summary['total_viewing_duration'] = result_row.Total_Hours
               summary['video_duration'] = result_row.V_Duration
               summary['average_view'] = result_row.Avgerage_Minutes             #TODO

               audience_profile_dict['Wirehouse Advisors'] = result_row.Wirehouse_Advisors
               audience_profile_dict['Independent B/D'] =result_row.Independent_BD
               audience_profile_dict['RIA'] = result_row.RIA
               audience_profile_dict['Insurance_CPAs_BankTrust'] = result_row.Insurance_CPAs_BankTrust
               audience_profile_dict['Investment Consultant'] = result_row.Investment_Consultant
               audience_profile_dict['Plan Sponsor'] = result_row.Plan_Sponsor
               audience_profile_dict['Endowment_Foundation']  =  result_row.Endowment_Foundation
               audience_profile_dict['Asset Manager'] = result_row.Asset_Manager
               audience_profile_dict['Other'] = result_row.Other

               # URL for Image and Video links
               url_video = result_row.V_VideoLink
               url = result_row.V_ImageURL

               #
               #get the video caption(s) and Company Name  from the database
               #
               caption_results = Video_Tag.query.filter_by(video_id = video_id )\
                   .order_by(Video_Tag.tag_name).all()

               caption = ''
               if caption_results:
                   for c in caption_results:
                       if c.tag_type =='People':
                           header['speaker'] = c.tag_name
                       if c.tag_type == 'Companies':
                           header['company_name'] =  c.tag_name

                   caption = 'PARTICIPANTS | ' + header['speaker'] + ' of ' + header['company_name']


               #
               # complete the report header
               #

               header['published_date'] = result_row.V_DatePublished.strftime('%B %d,%Y')
               #header['published_date'] = 'PUBLISHED | ' + result_row.V_DatePublished.strftime('%B %d,%Y')
               header['report_date'] = 'VIEWING REPORT | ' + str(current_month_number) + '/1/' + str(current_year)
               header['masterclass'] = masterclass
               header['single'] = single
               header['report_name'] = result_row.V_Title

               spreadsheet_name = result_row.V_Title + '_' + str(result_row.SPeriod) + '_' + str(result_row.SYear) + '.xlsm'

               #
               # top companies
               #

               top_companies =[]
               top_companies_result = TopCompany.query.filter_by(VTCVID = video_id).\
                   order_by(desc(TopCompany.VTCViews)).all()

               for index in  range(len(top_companies_result)):
                   if index == 10:
                       break
                   top_companies.append(top_companies_result[index].VTCCompany)

               #
               # get the data for the bar chart
               #
               barchart_data = []
               barchart_ticks = []

               for index in range(len(results)):

                   if results[index].month_id <= current_month_id:


                       # make ticks in the format of 1-JAN, 1-MAR , but look out for 1W
                       if  results[index].month_short_name[0].isdigit():   ## is the first character a digit
                           barchart_ticks.append([index,  results[index].month_short_name.encode("utf-8")])
                       else:
                           barchart_ticks.append([index, '1-' + results[index].month_short_name.encode("utf-8")])
                       barchart_data.append([index, results[index].total_views])


               try:
                   warnings.simplefilter("ignore")

                   workbook = load_workbook('spreadsheets/TEMPLATE_Non-Channel_online.xlsm',keep_vba=True)
                   ws = workbook.get_sheet_by_name("Source")
                   warnings.simplefilter("default")

                   ws['A2'] = 'SINGLE'
                   ws['B2'] = str(current_month_number) + '/1/' + str(current_year)
                   ws['C2']=  header['report_name'] # report title
                   ws['D2'] = header['published_date']
                   ws['E2'] = summary['video_duration']
                   ws['F2'] = header['company_name']  # participants
                   ws['G2'] = header['speaker']  # participants
                   ws['H2'] = url_video  # thumbnail image
                   ws['I2'] = url
                   ws['J2'] = summary['total_views']
                   ws['K2'] = summary['total_viewing_duration']
                   ws['L2'] = summary['average_view']



                   #top viewing companies
                   index = 5
                   rank = 1
                   for t in top_companies:  #top company data D17 to D26

                      ws['G' + str(index)] = t
                      ws['H' + str(index)] = rank
                      index = index +1
                      rank = rank + 1

                   # audience profile
                   ws['E5'] = audience_profile_dict['Wirehouse Advisors']
                   ws['E6'] = audience_profile_dict['Independent B/D']
                   ws['E7'] = audience_profile_dict['RIA']
                   ws['E8'] = audience_profile_dict['Insurance_CPAs_BankTrust']
                   ws['E9'] = audience_profile_dict['Investment Consultant']
                   ws['E10'] = audience_profile_dict['Plan Sponsor']
                   ws['E11'] = audience_profile_dict['Endowment_Foundation']
                   ws['E12'] = audience_profile_dict['Asset Manager']
                   ws['E13'] = audience_profile_dict['Other']

                   #views timeline
                   index = 5
                   for b in barchart_ticks:
                       ws['A' + str(index)] = b[1]
                       index = index + 1

                   index =5
                   for b in barchart_data:
                       ws['B' + str(index)] = b[1]
                       index = index + 1

                   #workbook.save(spreadsheet_name)
                   save_workbook(workbook,spreadsheet_name)
                   #return render_template('error.html',error_message='Good!')

                   return send_file(os.path.abspath('.') + '\\' + spreadsheet_name, as_attachment=True,
                                    attachment_filename=spreadsheet_name)


               except Exception as e:
                   print str(e)
                   return render_template('error.html',
                                          error_message='Unable to complete spreadsheet')



        return render_template('error.html',
                error_message='No data for the current reporting period, check database')


    else:
        #----------------------------------
        # masterclass report processing
        #----------------------------------

        result_row = None


        # loop through results set to find current month
        for row in results:
           if row.SPeriod == current_month and row.SYear == current_year :
             result_row = row
        if not result_row:
            return render_template('error.html',
                error_message='No data for the current reporting period, check database')




        # get viewing stats
        summary['total_viewing_duration'] =  0
        summary['total_views'] = 0
        summary['video_duration'] = result_row.V_Duration

        for result_row in results:
             if result_row.month_id <= current_month_id :
                summary['total_views'] = summary['total_views'] + result_row.total_views
                summary['total_viewing_duration'] = result_row.Total_Hours

        summary['average_view'] = summary['total_viewing_duration'] * 60 / summary['total_views']

        audience_profile_dict['Wirehouse Advisors'] = result_row.Wirehouse_Advisors
        audience_profile_dict['Independent B/D'] = result_row.Independent_BD
        audience_profile_dict['RIA'] = result_row.RIA
        audience_profile_dict['Insurance_CPAs_BankTrust'] = result_row.Insurance_CPAs_BankTrust
        audience_profile_dict['Investment Consultant'] = result_row.Investment_Consultant
        audience_profile_dict['Plan Sponsor'] = result_row.Plan_Sponsor
        audience_profile_dict['Endowment_Foundation'] = result_row.Endowment_Foundation
        audience_profile_dict['Asset Manager'] = result_row.Asset_Manager
        audience_profile_dict['Other'] = result_row.Other

        # these guys are NOT sums
        url_video = results[0].V_VideoLink
        url = results[0].V_ImageURL
        summary['video_duration'] = results[0].V_Duration

        # get the video caption(s) and Company Name  from the database
        caption_results = Video_Tag.query.filter_by(video_id=video_id) \
            .order_by(Video_Tag.tag_name).all()

        caption = ''
        if caption_results:
            for c in caption_results:
                if c.tag_type == 'Companies':
                    caption = caption + c.tag_name + ', '
            caption = caption.rstrip(', ')  # remove last comma

        header['published_date'] = results[0].V_DatePublished.strftime('%B %d,%Y')
        header['report_date'] = 'VIEWING REPORT | ' + str(current_month_number) + '/1/' + str(current_year)
        header['masterclass'] = masterclass
        header['single'] = single



        # chop MASTERCLASS TITLE
        temp = results[0].V_Title.strip('MASTERCLASS:')
        dash = temp.find('-', 1) - 1  # back up 1 position in fron of dash
        header['report_name'] = temp[1: dash]

        spreadsheet_name = header['report_name'] + '_' + str(results[0].SPeriod) + '_' + str(
            results[0].SYear) + '.xlsm'


        #
        # top companies
        #

        top_companies = []
        top_companies_result = db.session.query(Masterclass_Top_Companies).filter_by(VTCVID=video_id).\
            order_by(desc('VW_TOP_COMPANIES_MASTERCLASS.SUM_VIEWS')).all()

        #
        # get the data for the bar chart
        #
        for index in range(len(top_companies_result)):
            if index == 10 :
                break
            else:
                top_companies.append(top_companies_result[index].VTCCompany)

        #
        # get the data for the bar chart
        #
        barchart_ticks = []
        barchart_data = []
        running_sum = 0
        for index in range(len(results)):

            if results[index].month_id <= current_month_id:
                print results[index].month_id, current_month_id

                # make ticks in the format of 1-JAN, 1-MAR , but look out for 1W
                if results[index].month_short_name[0].isdigit():  ## is the first character a digit
                    barchart_ticks.append([index, results[index].month_short_name.encode("utf-8")])
                else:
                    barchart_ticks.append([index, '1-' + results[index].month_short_name.encode("utf-8")])

                running_sum = running_sum  + results[index].total_views
                barchart_data.append([index, running_sum])


        # get workbook


        try:
            warnings.simplefilter("ignore")
            workbook = load_workbook('spreadsheets/TEMPLATE_Non-Channel_online.xlsm',keep_vba=True)
            ws = workbook.get_sheet_by_name("Source")
            warnings.simplefilter("default")

            ws['A2'] = 'MASTERCLASS'
            ws['B2'] = str(current_month_number) + '/1/' + str(current_year)
            ws['C2'] = header['report_name']  # report title
            ws['D2'] = header['published_date']
            ws['E2'] = summary['video_duration']
            #ws['G2'] = header['company_name']  # participants
            ws['G2'] = caption  # participants
            ws['H2'] = url_video  # thumbnail image
            ws['I2'] = url
            ws['J2'] = summary['total_views']
            ws['K2'] = summary['total_viewing_duration']
            ws['L2'] = summary['average_view']


            # top viewing companies
            index = 5
            rank = 1
            for t in top_companies:  # top company data D17 to D26

                ws['G' + str(index)] = t
                ws['H' + str(index)] = rank
                index = index + 1
                rank = rank + 1

            # audience profile
            ws['E5'] = audience_profile_dict['Wirehouse Advisors']
            ws['E6'] = audience_profile_dict['Independent B/D']
            ws['E7'] = audience_profile_dict['RIA']
            ws['E8'] = audience_profile_dict['Insurance_CPAs_BankTrust']
            ws['E9'] = audience_profile_dict['Investment Consultant']
            ws['E10'] = audience_profile_dict['Plan Sponsor']
            ws['E11'] = audience_profile_dict['Endowment_Foundation']
            ws['E12'] = audience_profile_dict['Asset Manager']
            ws['E13'] = audience_profile_dict['Other']

            # views timeline
            index = 5
            for b in barchart_ticks:
                ws['A' + str(index)] = b[1]
                index = index + 1

            index = 5
            for b in barchart_data:
                ws['B' + str(index)] = b[1]
                index = index + 1

            workbook.save(spreadsheet_name)


            return send_file(os.path.abspath('.') + '\\' + spreadsheet_name,as_attachment = True,
                             attachment_filename=spreadsheet_name)


        except Exception as e:
            print str(e)
            return render_template('error.html',
                                   error_message='Unable to complete spreadsheet')

@app.route('/search/channel_report', methods=['GET','POST'])
def channel_report_search():

    results = Video.query.filter(Video.V_Type == 'CHANNEL') \
        .order_by(Video.V_Title).all()
    return render_template('channel_report_results.html',results=results)





# ------------------------------------------
#       channel reports - type 1
# ------------------------------------------

@app.route('/channel_report/1/<int:company_tag_id>')
def channel_report_1(company_tag_id):


    # get current month stats from database
    current_month_stats = Current_Month_Stats.query.all()

    # unable to retrieve the current reportng month/year, show error page
    if not current_month_stats:
        return render_template('error.html',
                               error_message='Unable to retrieve the current reportng month/year, check database')

    current_month = current_month_stats[0].month_name # March, April .....
    current_year =  current_month_stats[0].month_year # 2015, 2016 .....
    current_month_number  = current_month_stats[0].month_number # 1 = Jan, 3 = Mar......
    #current_month_id = current_month_stats[0].month_id # 14,15,16 primary key for Month_Reports table

    #
    # get the viewing stats for the channel reports
    #

    viewing_stats = db.session.query(Channel_Reports)\
        .filter_by(TL_TID=company_tag_id, SYear = current_year).all()


    if not viewing_stats:
        return render_template('error.html',
                        error_message='No data for the current reporting period, check database')

    page1 = db.session.query(Channel_Reports_Page1).filter_by(T_ID=company_tag_id).all()
    company_name = page1[0].T_TAG


    spreadsheet_name = 'Channel Report' + '_' + company_name + '_' + \
                       current_month + '_' + str (current_year) + '.xlsx'

    try:
        warnings.simplefilter("ignore")
        workbook = load_workbook('spreadsheets/Channel.xlsx')
        ws = workbook.get_sheet_by_name("Sheet1")
        warnings.simplefilter("default")

        ws['A2'] = 'CHANNEL'
        ws['B2'] = str(current_month_number) + '/1/' + str(current_year)
        ws['C2'] = company_name
        ws['D2'] = page1[0].Sum_Total_Views
        ws['E2'] = page1[0].Sum_Total_Hours
        ws['F2'] =  page1[0].Sum_Total_Hours * 60 / page1[0].Sum_Total_Views
        ws['H2'] = "Hey, I forgot the formula!"


        # get the top 2 videos
        top_videos = db.session.query(Channel_Reports.c.V_Title,func.sum(Channel_Reports.c.Total_Hours).label("total_hours"),
                            func.sum(Channel_Reports.c.Total_Views).label("total_views"),\
                            Channel_Reports.c.V_DatePublished.label("pub_date"))\
                            .filter(Channel_Reports.c.TL_TID==company_tag_id, Channel_Reports.c.SYear==current_year)\
                            .group_by(Channel_Reports.c.V_Title,Channel_Reports.c.V_DatePublished)\
                            .order_by(desc('total_views')).all()

        if top_videos[0]:
            #Most Watched Title
            ws['I2'] = top_videos[0].V_Title
            ws['J2'] = top_videos[0].total_views
            ws['K2'] = top_videos[0].total_hours

        if top_videos[1]:
            #Second Most Watched
            ws['L2'] = top_videos[1].V_Title
            ws['M2'] = top_videos[1].total_views
            ws['N2'] = top_videos[1].total_hours

        # get total new videos for this month
        total_new_videos_for_this_month  = 0
        for row in top_videos:
            if row.pub_date.month == current_month and row.pub_date.year == current_year:
                total_new_videos_for_this_month = total_new_videos_for_this_month + 1
        ws['G2'] = total_new_videos_for_this_month


        # get the top viewing month
        top_month = db.session.query(Channel_Reports.c.SPeriod,Channel_Reports.c.SYear,
                                      func.sum(Channel_Reports.c.Total_Hours).label("total_hours"),
                                      func.sum(Channel_Reports.c.Total_Views).label("total_views")
                                      ).filter(Channel_Reports.c.TL_TID == company_tag_id,
                                               Channel_Reports.c.SYear == current_year). \
            group_by(Channel_Reports.c.SPeriod,Channel_Reports.c.SYear). \
            order_by(desc('total_views')).all()

        if top_month:
            ws['O2'] = top_month[0].SPeriod + ' ' + str(top_month[0].SYear)
            ws['Q2'] = top_month[0].total_views
            ws['R2'] = top_month[0].total_hours



        # get previous month's id,year,name,etc.
        # month_id = 14
        # month_short_name = FEB
        # month_name = February
        # month_number = 2
        # month_year = 2016

        last_month =  db.session.query(Last_Month).all()

        # get total new videos for LAST month
        total_new_videos_for_last_month  = 0
        for row in top_videos:
            if row.pub_date.month == last_month[0].month_number and row.pub_date.year == last_month[0].month_year:
                total_new_videos_for_last_month = total_new_videos_for_last_month + 1
        ws['V2'] = total_new_videos_for_last_month

        # get stats for LAST month
        last_month_stats  = db.session.query(Channel_Reports_Last_Month).filter_by(T_ID =company_tag_id ).all()

        if last_month_stats:
            ws['S2'] = last_month_stats[0].Sum_Total_Views
            ws['T2'] = last_month_stats[0].Sum_Total_Hours
            ws['U2'] = last_month_stats[0].Sum_Total_Hours * 60 / last_month_stats[0].Sum_Total_Views
            ws['H17']= last_month_stats[0].Wirehouse_Advisors
            ws['H18'] = last_month_stats[0].Independent_BD
            ws['H19'] = last_month_stats[0].RIA
            ws['H20'] = last_month_stats[0].Insurance_CPAs_BankTrust
            ws['H21'] = last_month_stats[0].Investment_Consultant
            ws['H22'] = last_month_stats[0].Plan_Sponsor
            ws['H23'] = last_month_stats[0].Endowment_Foundation
            ws['H24'] = last_month_stats[0].Asset_Manager
            ws['H25'] = last_month_stats[0].Other



        #ytd summary
        # get the top 2 videos
        ytd_summary = db.session.query(month_table.c.month,Channel_Reports.c.SPeriod,
                            func.sum(Channel_Reports.c.Total_Hours).label("total_hours"),
                            func.sum(Channel_Reports.c.Total_Views).label("total_views")
                            ).filter(Channel_Reports.c.TL_TID==company_tag_id, Channel_Reports.c.SYear==current_year)\
                            .join(Channel_Reports, month_table.c.month_name == Channel_Reports.c.SPeriod) \
                            .group_by(Channel_Reports.c.SPeriod,month_table.c.month). \
                            order_by('month').all()

        row_pointer = 5
        for row in ytd_summary:
            ws['A' +str(row_pointer)] = row.SPeriod
            ws['C' +str(row_pointer)] = row.total_views
            ws['D' +str(row_pointer)] = row.total_hours
            ws['E' + str(row_pointer)] = row.total_hours * 60 / row.total_views
            row_pointer = row_pointer + 1

        #get YTD Audience Category
        ytd_audience = db.session.query(CHANNEL_YTD_Audience_Category)\
            .filter(CHANNEL_YTD_Audience_Category.c.T_ID==company_tag_id).all()
        if ytd_audience:
            ws['H5'] = ytd_audience[0].Wirehouse_Advisors
            ws['H6'] = ytd_audience[0].Independent_BD
            ws['H7'] = ytd_audience[0].RIA
            ws['H8'] = ytd_audience[0].Insurance_CPAs_BankTrust
            ws['H9'] = ytd_audience[0].Investment_Consultant
            ws['H10'] = ytd_audience[0].Plan_Sponsor
            ws['H11'] = ytd_audience[0].Endowment_Foundation
            ws['H12'] = ytd_audience[0].Asset_Manager
            ws['H13'] = ytd_audience[0].Other

        ytd_matrix = db.session.query(CHANNEL_YTD_Reports)\
            .filter(CHANNEL_YTD_Reports.c.T_ID==company_tag_id).all()


        if ytd_matrix:
            row_pointer = 29
            for row in ytd_matrix:
                ws['A' +str(row_pointer)] = row.V_Title
                ws['B' +str(row_pointer)] = row.V_DatePublished
                ws['C' + str(row_pointer)] = row.V_Duration
                ws['D' +str(row_pointer)] = row.YTD_Views
                ws['E' + str(row_pointer)] = row.YTD_Total_Hours
                ws['F' + str(row_pointer)] = row.YTD_Avg_Min
                ws['G' + str(row_pointer)] = 'IDK'
                ws['H' + str(row_pointer)] = 'IDK'
                ws['I' + str(row_pointer)] = 'IDK'
                ws['J' + str(row_pointer)] = row.YTD_Wirehouse_Advisors
                ws['K' + str(row_pointer)] = row.YTD_Independent_BD
                ws['L' + str(row_pointer)] = row.YTD_RIA
                ws['M' + str(row_pointer)] = row.YTD_Investment_Consultant
                ws['N' +str(row_pointer)] = row.YTD_Plan_Sponsor
                ws['O' + str(row_pointer)] = row.YTD_Asset_Manager
                ws['P' + str(row_pointer)] = row.YTD_Other

                row_pointer = row_pointer + 1


        #get last month datail
        last_month_detail = db.session.query(Channel_last_month_Detail) \
            .filter(Channel_last_month_Detail.c.T_ID == company_tag_id).all()

        if last_month_detail:
            row_pointer = 29
            for row in last_month_detail:
                ws['R' + str(row_pointer)] = row.V_Title
                ws['S' + str(row_pointer)] = row.V_DatePublished
                ws['T' + str(row_pointer)] = row.V_Duration
                ws['U' + str(row_pointer)] = row.Total_Views
                ws['V' + str(row_pointer)] = row.Total_Hours
                ws['W' + str(row_pointer)] = 'IDK'
                ws['X' + str(row_pointer)] =  'IDK'
                ws['Y' + str(row_pointer)] = 'IDK'
                ws['Z' + str(row_pointer)] = 'IDK'
                ws['AA' + str(row_pointer)] = row.Wirehouse_Advisors
                ws['AB' + str(row_pointer)] = row.Independent_BD
                ws['AC' + str(row_pointer)] = row.RIA
                ws['AD' + str(row_pointer)] = row.Investment_Consultant
                ws['AE' + str(row_pointer)] = row.Plan_Sponsor
                ws['AF' + str(row_pointer)] = row.Asset_Manager
                ws['AG' + str(row_pointer)] = row.Other

                row_pointer = row_pointer + 1


        # get all time stats
        all_time_stats = db.session.query(channel_report_all_time_2)\
            .filter(channel_report_all_time_2.c.T_ID == company_tag_id).all()
        if all_time_stats:

             row_pointer = 29
             for row in all_time_stats:
                 ws['AI' + str(row_pointer)] = row.V_Title
                 ws['AJ' + str(row_pointer)] = row.V_DatePublished
                 ws['AK' + str(row_pointer)] = row.V_Duration
                 ws['AL' + str(row_pointer)] = row.Total_Views
                 ws['AM' + str(row_pointer)] = row.Total_Hours
                 ws['AN' + str(row_pointer)] = row.Average_Minutes
                 ws['AO' + str(row_pointer)] = row.Engagement

                 row_pointer = row_pointer + 1


        # get email stats #TODO
        email_stats = db.session.query(channel_email_stats).filter(channel_email_stats.c.ECID==10021) #TODO
        if email_stats:
            row_pointer = 5
            for row in email_stats:
                ws['M' + str(row_pointer)] = row.ETitle
                ws['N' + str(row_pointer)] = row.EDate
                ws['O' + str(row_pointer)] = row.ESends
                ws['P' + str(row_pointer)] = row.EOpens
                ws['Q' + str(row_pointer)] = row.ECTR
                ws['R' + str(row_pointer)] = row.EWirehouseAdvisors
                ws['S' + str(row_pointer)] = row.EIndepentBD
                ws['T' + str(row_pointer)] = row.ERIA
                ws['U' + str(row_pointer)] = row.EInvestmentConsultant
                ws['V' + str(row_pointer)] = row.EPlanSponsor
                ws['W' + str(row_pointer)] = row.EAssetManager
                ws['X' + str(row_pointer)] = row.EOther
                row_pointer = row_pointer + 1



        workbook.save(spreadsheet_name)

        return send_file(os.path.abspath('.') + '\\' + spreadsheet_name, as_attachment=True,
                         attachment_filename=spreadsheet_name)


    except Exception as e:
        print str(e)
        return render_template('error.html',
                               error_message='Unable to complete spreadsheet')

